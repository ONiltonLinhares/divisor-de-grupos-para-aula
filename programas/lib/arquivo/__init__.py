from lib.interface import *
import openpyxl

def arquivoexiste(nome):
    try:
        a = open(nome, 'rt')
        a.close()
    except FileNotFoundError:
        return False
    else:
        return True

def planilhaexiste(nome):
    try:
        book = openpyxl.load_workbook(nome)
    except FileNotFoundError:
        return False
    else:
        return True

def craiarquivo(nome):
    try:
        a = open(nome, 'wt+')
        a.close()
    except:
        print('\033[31mERRO: Houve um problema na criação do arquivo.\033[m')

def craiarplanilha(nome, pagina = 'pagina-1'):
    try:
        book = openpyxl.Workbook()
        book.save(nome)
    except:
        print('\033[31mERRO: Houve um problema na criação da planilha.\033[m')
    try:
        book1 = book.active
        book1.title = pagina
        book.save(nome)
    except:
        print('\033[31mERRO: Houve um problema na mudança de nome da pagina 1.\033[m')

def lerarquivo(nome, str):
    try:
        a = open(nome, 'rt')
    except:
        print('\033[31mERRO: Problema ao ler o arquivo\033[m')
    else: 
        cabeçalho(str)
        print('\033[33m-\033[m' * 88)
        for linha in a:
            dado = linha.split(';')
            dado[3] = dado[3].replace('\n', '')
            print(f'\033[32m{dado[0]:<34}{dado[1]:>3} Anos |CPF:{dado[2]:<8} |Celular:{dado[3]:>13}\033[m')
        print('\033[33m-\033[m' * 88)
    finally:
        a.close()

def lerplanilha(nome, nomepagina, str = 'PLANILHA'):
    try:
        planilha = openpyxl.load_workbook(nome)
        pagina = planilha[nomepagina]
    except:
        print('\033[31mERRO: Problema ao ler a planilha\033[m')
    else:
        cabeçalho(str)
        print('\033[33m-\033[m' * 88)
        for linhas in pagina:
            print(f'\033[32m{linhas[0].value:<34}{linhas[1].value:>3} Anos |CPF:{linhas[2].value:<8} |Celular:{linhas[3].value:>13}\033[m')
        print('\033[33m-\033[m' * 88)


def cadastrartxt(arq, nome = 'DESCONHECIDO', idade = 0, cpf = 'NÃO IMFORMADO', telefone = 'SEM CONTATO'):
    try:
        a = open(arq, 'at')
    except:
        print('\033[31mHouve um ERRO na abertura do arquivo\033[m')
    else:
        try:
            a.write(f'{nome};{idade};{cpf};{telefone}\n')
        except:
            print('\033[31mHouve um ERRO na escrita dos dados\033[m')
        else:
            print(f'Novo registro de {nome} adicionado.')
            a.close()

def escrever_txt (arq, txt):
    try:
        a = open(arq, 'at')
    except:
        print('\033[31mHouve um ERRO na abertura do arquivo\033[m')
    else:
        try:
            a.write(txt)
        except:
            print('\033[31mHouve um ERRO na escrita dos dados\033[m')



def cadastrar(arq, page, nome = 'DESCONHECIDO', idade = 0, cpf = 'NÃO IMFORMADO', telefone = 'SEM CONTATO'):
    try:
        planilha = openpyxl.load_workbook(arq)
        pagina = planilha[page]
    except:
        print('\033[31mHouve um ERRO na abertura da planilha\033[m')
    else:
        try:
            pagina.append([nome, idade, cpf, telefone])
            planilha.save(arq)
        except:
            print('\033[31mHouve um ERRO na escrita dos dados\033[m')
        else:
            cabeçalho(f'Novo registro de {nome} adicionado.')

def acharuser():
    a = open('usuarios2.txt', 'wt')
    dados = a.readlines()
    marcador = False
    userintermediario = []
    user =  []
    while user not in dados:
        txt = input('Digite o nome(como consta no cadastro) de um usuario:')
        for i1 in dados:
            userintermediario = i1.split(';')
            if userintermediario[0] == txt:
                user = userintermediario
                marcador = True
        if marcador == False:
            print('\033[31mERRO: Usuario não encontrado.\033[m')
        else:
            return user

def excluiusertxt():
    a = open('usuarios2.txt', 'rt')
    user = acharuser()
    dados = a.readlines()
    novos_dados = []
    a.close()
    b = open('usuarios2.txt', 'wt')
    marcador = 0
    try:
        for i in dados:
            if dados[marcador].split(';') != user:
                    b.write(dados[marcador])
            marcador += 1
    except:
        print('\033[31mERRO: Problema ao tentar excluir o usuario.\033[m')
    else:
        cabeçalho('USUARIO EXCLUIDO', 'vm')

def excluiuser(nome, nomepagina):
    txt = input('Digite o nome(como consta no cadastro) de um usuario:')
    planilha = openpyxl.load_workbook(nome)
    pagina = planilha[nomepagina]
    user = False
    for linha in pagina:
        for celula in linha:
            if celula.value == txt:
                user = True
            if user == True:
              pagina.delete_rows(celula.row)
              planilha.save('usuarios2.xlsx')
                


