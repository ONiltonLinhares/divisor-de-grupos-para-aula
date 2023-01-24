import openpyxl

# Open the Excel file and get the worksheet
workbook = openpyxl.load_workbook('people (1).xlsx')
worksheet = workbook['Sheet1']

# Get the list of names from the worksheet
names = []
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    name = row[0]
    names.append(name)

# Calculate the number of people in each main group
num_people = len(names)
num_people_per_group = num_people // 2

# Divide the names into the main groups and subgroups
group_a = names[:num_people_per_group]
group_b = names[num_people_per_group:]

# Divide each main group into subgroups of four
subgroups_a = [group_a[i:i+4] for i in range(0, len(group_a), 4)]
subgroups_b = [group_b[i:i+4] for i in range(0, len(group_b), 4)]

# Print the main group A and its subgroups
print("Group A:")
for i, subgroup in enumerate(subgroups_a):
    print(f"Subgroup {i+1}: {', '.join(subgroup)}")

# Print the main group B and its subgroups
print("\nGroup B:")
for i, subgroup in enumerate(subgroups_b):
    print(f"Subgroup {i+1}: {', '.join(subgroup)}")

