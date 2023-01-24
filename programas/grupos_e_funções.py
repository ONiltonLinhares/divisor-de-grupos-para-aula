from lib.arquivo import *
import openpyxl
import random

# creating the last file
arq = 'grupos_teste1.txt'
arq2 = 'grupos_teste2.txt'
arq3 = 'grupos_teste_geral.txt'

if not arquivoexiste(arq):
    craiarquivo(arq)

if not arquivoexiste(arq2):
    craiarquivo(arq2)

if not arquivoexiste(arq2):
    craiarquivo(arq2)

# Open the Excel file and get the worksheet
workbook = openpyxl.load_workbook('people (1).xlsx')
worksheet = workbook['Sheet1']

# Get the list of names from the worksheet
names = []
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    name = row[0]
    names.append(name)

# Shuffle the names randomly
random.shuffle(names)

# Calculate the number of people in each main group
num_people = len(names)
num_people_per_group = num_people // 2

# Divide the names into the main groups and subgroups
group_a = names[:num_people_per_group]
group_b = names[num_people_per_group:]

# Divide each main group into subgroups of four
subgroups_a = [group_a[i:i+4] for i in range(0, len(group_a), 4)]
subgroups_b = [group_b[i:i+4] for i in range(0, len(group_b), 4)]

# Save the group 1 in a txt doc 
a = open(arq, 'wt')
a.write(f'{subgroups_a}')

# Save the group 2 in a txt doc 
a2 = open(arq2, 'wt')
a2.write(f'{subgroups_b}')

# Save the main groups in a txt doc
a3 = open(arq3, 'wt')
a3.write('Grupo 1:\n')
for people in group_a:
    a3.write(f'\n{people}')

a3.write('\n\nGrupo 2:\n')
for people in group_b:
    a3.write(f'\n{people}')